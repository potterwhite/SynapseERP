# Copyright (c) 2025-present, PotterWhite (themanuknowwhom@outlook.com).
# All rights reserved.
#
# This source code is licensed under the MIT license found in the
# LICENSE file in the root directory of this source tree.
#
# T-HEAD-GR-V0.1.0-20250905 (Engine writer module for BOM Analyzer)

import io
import pandas as pd
from openpyxl.styles import Alignment, PatternFill
from django.utils.translation import gettext_lazy as _


def func_3_0_format_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    df_formatted = dataframe.copy()

    # --- Sorting Logic (remains the same, it's correct) ---
    sort_keys = [col for col in ["值", "封装", "备注"] if col in df_formatted.columns]
    if "is_suspicious" in df_formatted.columns:
        primary_sort_col = "is_suspicious"
        sort_by_cols = [primary_sort_col] + sort_keys
        sort_ascending_order = [False] + [True] * len(sort_keys)
        df_formatted = df_formatted.sort_values(
            by=sort_by_cols, ascending=sort_ascending_order
        ).reset_index(drop=True)
    # --- End of Sorting Logic ---

    df_formatted.insert(0, "序号", range(1, len(df_formatted) + 1))

    rename_map = {
        "quantity_details": str(_("Quantity Breakdown")),
        "total_quantity": str(_("Total Quantity")),
    }
    df_formatted.rename(columns=rename_map, inplace=True)

    desired_column_order = [
        "序号",
        "值",
        "封装",
        "备注",
        str(_("Quantity Breakdown")),
        str(_("Total Quantity")),
    ]

    output_columns = [
        col for col in desired_column_order if col in df_formatted.columns
    ]
    if "is_suspicious" in df_formatted.columns:
        output_columns.append("is_suspicious")

    df_formatted = df_formatted[output_columns]

    return df_formatted


def func_2_2_generate_excel_output(final_dataframe: pd.DataFrame) -> io.BytesIO:
    """
    Generates the final Excel file, with suspicious rows sorted to the top AND highlighted.
    """
    # --- KEY CHANGE: We format the DataFrame FIRST, while 'is_suspicious' still exists. ---
    df_formatted = func_3_0_format_dataframe(final_dataframe)

    # Now, we prepare a version for writing to Excel by dropping the internal flag column.
    df_to_write = df_formatted.drop(columns=["is_suspicious"], errors="ignore")
    # --- END OF KEY CHANGE ---

    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
        df_to_write.to_excel(writer, sheet_name="Aggregated BOM", index=False)

        worksheet = writer.sheets["Aggregated BOM"]
        wrap_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

        # --- NEW: STYLING LOGIC to add color highlighting ---
        # Define the fill color for suspicious rows (a light red)
        suspicious_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )

        # Iterate through the original FORMATTED dataframe to check the flag
        for idx, row in df_formatted.iterrows():
            if row["is_suspicious"]:
                # The row index in the worksheet is idx + 2 (1 for header, 1 for 0-based to 1-based index)
                for col_idx in range(1, len(df_to_write.columns) + 1):
                    worksheet.cell(row=idx + 2, column=col_idx).fill = suspicious_fill
        # --- END OF NEW STYLING LOGIC ---

        # Apply text wrapping to the breakdown column (as before)
        column_to_wrap = str(_("Quantity Breakdown"))
        for col_idx, col_name in enumerate(df_to_write.columns, 1):
            if col_name == column_to_wrap:
                col_letter = chr(ord("A") + col_idx - 1)
                worksheet.column_dimensions[col_letter].width = 45
                for cell in worksheet[col_letter]:
                    cell.alignment = wrap_alignment
                break

    output_buffer.seek(0)
    return output_buffer
